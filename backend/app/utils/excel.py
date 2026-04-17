from __future__ import annotations

from io import BytesIO
from typing import Sequence

import pandas as pd
from openpyxl.utils import get_column_letter

from app.models.fred_models import FredSeriesRow
from app.models.request_models import GridObservation, Observation
from app.models.worldbank_models import WorldBankRow


ExcelCellValue = str | int | float | None


def _measure_width(values: Sequence[ExcelCellValue]) -> int:
    return min(max(max(len(str(value)) for value in values), 12) + 2, 42)


def _sanitize_file_segment(value: str) -> str:
    sanitized = []
    for character in value:
        if character in '<>:"/\\|?*' or ord(character) < 32:
            sanitized.append(" ")
        else:
            sanitized.append(character)
    return "_".join(" ".join("".join(sanitized).split()).split())


def _normalize_excel_value(value: ExcelCellValue) -> str | int | float:
    if value in (None, ""):
        return ""
    return value


def _build_flat_workbook(
    *,
    records: list[dict[str, ExcelCellValue]],
    columns: Sequence[str],
    file_name: str,
    sheet_name: str,
) -> tuple[BytesIO, str]:
    if not records:
        raise ValueError("No data available.")

    normalized_records = [{column: _normalize_excel_value(record.get(column)) for column in columns} for record in records]
    buffer = BytesIO()
    dataframe = pd.DataFrame(normalized_records, columns=list(columns))

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        dataframe.to_excel(writer, sheet_name=sheet_name, index=False)

        worksheet = writer.sheets[sheet_name]
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(normalized_records) + 1}"

        for column_index, column_name in enumerate(columns, start=1):
            worksheet.column_dimensions[get_column_letter(column_index)].width = _measure_width(
                [column_name, *[record[column_name] for record in normalized_records]]
            )

    buffer.seek(0)
    return buffer, file_name


def build_excel_workbook(country: str, indicator: str, observations: Sequence[Observation]) -> tuple[BytesIO, str]:
    if not observations:
        raise ValueError("No data available.")

    buffer = BytesIO()
    rows = [
        {
            "Year": observation.year,
            "Value": _normalize_excel_value(observation.value),
        }
        for observation in sorted(observations, key=lambda row: row.year)
    ]
    dataframe = pd.DataFrame(rows, columns=["Year", "Value"])

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        metadata_frame = pd.DataFrame([["Country", country], ["Indicator", indicator]])
        metadata_frame.to_excel(writer, sheet_name="IMF Data", index=False, header=False)
        dataframe.to_excel(writer, sheet_name="IMF Data", index=False, startrow=3)

        worksheet = writer.sheets["IMF Data"]
        worksheet.freeze_panes = "A5"
        worksheet.auto_filter.ref = f"A4:B{len(rows) + 4}"
        worksheet.column_dimensions[get_column_letter(1)].width = _measure_width(["Year", *[entry["Year"] for entry in rows]])
        worksheet.column_dimensions[get_column_letter(2)].width = _measure_width(["Value", *[entry["Value"] for entry in rows]])

    buffer.seek(0)
    file_name = f"{_sanitize_file_segment(country)}_{_sanitize_file_segment(indicator)}.xlsx"
    return buffer, file_name


def build_imf_grid_workbook(rows: Sequence[GridObservation]) -> tuple[BytesIO, str]:
    records = [
        {
            "Country": row.country,
            "Indicator": row.indicator,
            "Year": row.year,
            "Value": row.value,
        }
        for row in sorted(rows, key=lambda item: (item.country.casefold(), item.indicator.casefold(), item.year))
    ]
    return _build_flat_workbook(
        records=records,
        columns=["Country", "Indicator", "Year", "Value"],
        file_name=f"imf_data_grid_{pd.Timestamp.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx",
        sheet_name="IMF Data",
    )


def build_world_bank_workbook(rows: Sequence[WorldBankRow]) -> tuple[BytesIO, str]:
    records = [
        {
            "Country": row.country,
            "Indicator": row.indicator,
            "Year": row.year,
            "Value": row.value,
        }
        for row in sorted(rows, key=lambda item: (item.country.casefold(), item.indicator.casefold(), item.year))
    ]
    return _build_flat_workbook(
        records=records,
        columns=["Country", "Indicator", "Year", "Value"],
        file_name=f"world_bank_data_{pd.Timestamp.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx",
        sheet_name="World Bank Data",
    )


def build_fred_workbook(rows: Sequence[FredSeriesRow]) -> tuple[BytesIO, str]:
    records = [
        {
            "Series ID": row.series_id,
            "Title": row.title,
            "Year": row.date,
            "Value": row.value,
        }
        for row in sorted(rows, key=lambda item: (item.title.casefold(), item.series_id.casefold(), int(item.date)))
    ]
    return _build_flat_workbook(
        records=records,
        columns=["Series ID", "Title", "Year", "Value"],
        file_name=f"fred_data_{pd.Timestamp.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx",
        sheet_name="FRED Data",
    )
